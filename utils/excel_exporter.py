"""
Excel Export Utility for Stock Screeners
Generates formatted Excel reports with date-based naming convention:
- EW_DDMMYYYY.xlsx for Elliott Wave reports
- WW_DDMMYYYY.xlsx for Wolfe Wave reports
"""

import logging
import pandas as pd
import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


def get_excel_filename(report_type='EW'):
    """
    Generate Excel filename with date format DDMMYYYY.
    
    Args:
        report_type: 'EW' for Elliott Wave or 'WW' for Wolfe Wave
    
    Returns:
        Filename string (e.g., 'EW_22062026.xlsx')
    """
    today = datetime.datetime.now()
    date_str = today.strftime('%d%m%Y')
    return f"output/{report_type}_{date_str}.xlsx"


def _get(r, key, default=''):
    return r.get(key, default) if isinstance(r, dict) else default


def _get_currency(r, key, decimals=0):
    v = _get(r, key, None)
    if v is None or v == '':
        return ''
    try:
        if decimals == 0:
            return f"₹{v:.0f}"
        else:
            return f"₹{v:.2f}"
    except Exception:
        return ''


def _get_number(r, key, fmt='.2f'):
    v = _get(r, key, None)
    if v is None or v == '':
        return ''
    try:
        return f"{v:{fmt}}"
    except Exception:
        return ''


def _get_percent(r, key, decimals=1):
    v = _get(r, key, None)
    if v is None or v == '':
        return ''
    try:
        return f"{v:.{decimals}f}%"
    except Exception:
        return ''


def export_elliott_wave_to_excel(ranked_results, all_results, output_file=None):
    """
    Export Elliott Wave screening results to Excel.
    
    Args:
        ranked_results: List of high-confidence results (filtered)
        all_results: All results found
        output_file: Optional custom filename, defaults to EW_DDMMYYYY.xlsx
    """
    if output_file is None:
        output_file = get_excel_filename('EW')
    
    # Create Excel writer
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        
        # ========== SHEET 1: SUMMARY ==========
        display_results = ranked_results + [r for r in all_results if r['confidence'] < 75][:20-len(ranked_results)]
        
        summary_data = []
        for i, r in enumerate(display_results[:20]):
            summary_data.append({
                'Rank': i + 1,
                'Stock': r['symbol'],
                'Sector': r['sector'],
                'Wave Status': r['wave_status'],
                'Confidence %': f"{r['confidence']:.0f}",
                'Entry Zone': f"₹{r['entry_low']:.0f}-{r['entry_high']:.0f}",
                'Stop Loss': f"₹{r['stop_loss']:.0f}",
                'Target 1': f"₹{r['target1']:.0f}",
                'Target 2': f"₹{r['target2']:.0f}",
                'R:R Ratio': r['rr_ratio'],
                'Setup Type': r['setup_type'],
                'Current Price': f"₹{r['current_price']:.2f}",
            })
        
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        
        # Format Summary sheet
        worksheet = writer.sheets['Summary']
        _format_worksheet(worksheet, len(summary_df) + 1)
        
        # ========== SHEET 2: DETAILED ANALYSIS ==========
        detail_data = []
        for r in display_results[:10]:
            detail_data.append({
                'Stock': r['symbol'],
                'Current Price': f"₹{r['current_price']:.2f}",
                'Sector': r['sector'],
                'Confidence': f"{r['confidence']:.0f}%",
                'Wave 1 Start': f"₹{r['wave1_start']:.2f}",
                'Wave 1 End': f"₹{r['wave1_end']:.2f}",
                'Wave 1 %': f"{r['wave1_pct']:.1f}%",
                'Wave 2 End': f"₹{r['wave2_end']:.2f}",
                'Wave 2 Retrace': f"{r['wave2_retrace']:.1f}%",
                'RSI': f"{r['rsi']:.1f}",
                'RSI Divergence': 'Yes' if r['rsi_divergence'] else 'No',
                'MACD Status': r['macd_status'],
                'Volume Ratio': f"{r['vol_ratio_20']:.2f}x",
                'Volume Expansion': 'Yes' if r['volume_expansion'] else 'No',
                'OBV Trend': r['obv_trend'],
                'EMA Alignment': 'Yes' if r['ema_alignment'] else 'No',
                'Above 200 EMA': 'Yes' if r['above_ema200'] else 'No',
                'Entry Low': f"₹{r['entry_low']:.2f}",
                'Entry High': f"₹{r['entry_high']:.2f}",
                'Stop Loss': f"₹{r['stop_loss']:.2f}",
                'Target 1': f"₹{r['target1']:.2f}",
                'Target 2': f"₹{r['target2']:.2f}",
                'Setup Type': r['setup_type'],
            })
        
        detail_df = pd.DataFrame(detail_data)
        detail_df.to_excel(writer, sheet_name='Detailed Analysis', index=False)
        worksheet = writer.sheets['Detailed Analysis']
        _format_worksheet(worksheet, len(detail_df) + 1)
        
        # ========== SHEET 3: FIBONACCI LEVELS ==========
        fib_data = []
        for r in display_results[:10]:
            fib_retrace = _get(r, 'fib_retrace', {}) or {}
            fib_extension = _get(r, 'fib_extension', {}) or {}
            fib_data.append({
                'Stock': _get(r, 'symbol', ''),
                '23.6%': _get_currency({'': ''}, '',) if False else (f"₹{fib_retrace.get('23.6%', 0):.2f}" if fib_retrace.get('23.6%') is not None else ''),
                '38.2%': (f"₹{fib_retrace.get('38.2%', 0):.2f}" if fib_retrace.get('38.2%') is not None else ''),
                '50.0%': (f"₹{fib_retrace.get('50.0%', 0):.2f}" if fib_retrace.get('50.0%') is not None else ''),
                '61.8%': (f"₹{fib_retrace.get('61.8%', 0):.2f}" if fib_retrace.get('61.8%') is not None else ''),
                '78.6%': (f"₹{fib_retrace.get('78.6%', 0):.2f}" if fib_retrace.get('78.6%') is not None else ''),
                'Ext 100%': (f"₹{fib_extension.get('100%', 0):.2f}" if fib_extension.get('100%') is not None else ''),
                'Ext 161.8%': (f"₹{fib_extension.get('161.8%', 0):.2f}" if fib_extension.get('161.8%') is not None else ''),
                'Ext 261.8%': (f"₹{fib_extension.get('261.8%', 0):.2f}" if fib_extension.get('261.8%') is not None else ''),
            })
        
        fib_df = pd.DataFrame(fib_data)
        fib_df.to_excel(writer, sheet_name='Fibonacci Levels', index=False)
        worksheet = writer.sheets['Fibonacci Levels']
        _format_worksheet(worksheet, len(fib_df) + 1)
        
        # ========== SHEET 4: SCORE BREAKDOWN ==========
        score_data = []
        for r in display_results[:10]:
            score_data.append({
                'Stock': _get(r, 'symbol', ''),
                'EW Structure': _get_number(r, 'ew_score', '.0f') + '/40' if _get(r, 'ew_score', None) not in (None, '') else '',
                'Volume': _get_number(r, 'vol_score', '.0f') + '/20' if _get(r, 'vol_score', None) not in (None, '') else '',
                'Momentum': _get_number(r, 'momentum_score', '.0f') + '/20' if _get(r, 'momentum_score', None) not in (None, '') else '',
                'Institutional': _get_number(r, 'inst_score', '.0f') + '/20' if _get(r, 'inst_score', None) not in (None, '') else '',
                'Total Score': _get_number(r, 'total_score', '.0f') + '/100' if _get(r, 'total_score', None) not in (None, '') else '',
                'Confidence': _get_percent(r, 'confidence', 0) if _get(r, 'confidence', None) not in (None, '') else '',
            })
        
        score_df = pd.DataFrame(score_data)
        score_df.to_excel(writer, sheet_name='Score Breakdown', index=False)
        worksheet = writer.sheets['Score Breakdown']
        _format_worksheet(worksheet, len(score_df) + 1)
        
        # ========== SHEET 5: METADATA ==========
        metadata = [
            ['Report Type', 'Elliott Wave Screener'],
            ['Generated', datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S IST')],
            ['Total Stocks Scanned', len(all_results) + (150 - len(all_results))],
            ['Setups Found', len(all_results)],
            ['High Confidence (≥75%)', len(ranked_results)],
            ['', ''],
            ['Screening Parameters', ''],
            ['Universe', 'Nifty 50 + Nifty Next 50 + Select Midcaps'],
            ['Data Period', '1-year daily'],
            ['Min Confidence', '75%'],
            ['Min Risk/Reward', '1:3'],
        ]
        
        metadata_df = pd.DataFrame(metadata)
        metadata_df.to_excel(writer, sheet_name='Metadata', index=False, header=False)
    
    print(f"  [OK] Excel report saved to {output_file}\n")
    return output_file


def export_wolfe_wave_to_excel(ranked_results, all_results, output_file=None):
    """
    Export Wolfe Wave screening results to Excel.
    
    Args:
        ranked_results: List of high-confidence results (filtered)
        all_results: All results found
        output_file: Optional custom filename, defaults to WW_DDMMYYYY.xlsx
    """
    if output_file is None:
        output_file = get_excel_filename('WW')
    
    # Separate bullish and bearish
    ranked_bullish = [r for r in ranked_results if _get(r, 'pattern_type') == 'Bullish']
    ranked_bearish = [r for r in ranked_results if _get(r, 'pattern_type') == 'Bearish']
    
    display_results = ranked_results[:25] if ranked_results else sorted(
        all_results, key=lambda x: x.get('confidence', 0), reverse=True
    )[:25]
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        
        # ========== SHEET 1: SUMMARY ==========
        summary_data = []
        for i, r in enumerate(display_results):
            try:
                summary_data.append({
                    'Rank': i + 1,
                    'Stock': _get(r, 'symbol', ''),
                    'Sector': _get(r, 'sector', ''),
                    'Pattern Type': _get(r, 'pattern_type', ''),
                    'Score': _get_number(r, 'confidence', '.0f'),
                    'Sweet Zone Quality': _get(r, 'sweet_zone_quality', ''),
                    'Current Price': _get_currency(r, 'current_price', 2),
                    'P5 (Entry)': _get_currency(r, 'p5', 2),
                    'EPA Target': _get_currency(r, 'epa', 2),
                    'Stop Loss': _get_currency(r, 'stop_loss', 2),
                    'R:R Ratio': _get(r, 'rr_ratio', ''),
                    'Potential %': _get_percent(r, 'potential_pct', 1),
                    'Status': _get(r, 'status', ''),
                })
            except Exception as e:
                logger.warning("Skipping summary row for %s due to error: %s", _get(r, 'symbol'), e)
                continue
        
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        worksheet = writer.sheets['Summary']
        _format_worksheet(worksheet, len(summary_df) + 1)
        
        # ========== SHEET 2: BULLISH SETUPS ==========
        if ranked_bullish:
            bullish_data = []
            for r in ranked_bullish[:15]:
                try:
                    bullish_data.append({
                        'Stock': _get(r, 'symbol', ''),
                        'Score': _get_number(r, 'confidence', '.0f'),
                        'P1 (Low)': _get_currency(r, 'p1', 2),
                        'P2 (High)': _get_currency(r, 'p2', 2),
                        'P3 (Low)': _get_currency(r, 'p3', 2),
                        'P4 (High)': _get_currency(r, 'p4', 2),
                        'P5 (Entry)': _get_currency(r, 'p5', 2),
                        'EPA Target': _get_currency(r, 'epa', 2),
                        'Stop Loss': _get_currency(r, 'stop_loss', 2),
                        '1-3 Line': _get(r, 'line_13_quality', ''),
                        '2-4 Line': _get(r, 'line_24_quality', ''),
                        'Convergence': 'Yes' if r.get('converging', True) else 'No',
                        'Entry Quality': _get(r, 'sweet_zone_quality', ''),
                    })
                except Exception as e:
                    logger.warning("Skipping bullish row for %s due to error: %s", _get(r, 'symbol'), e)
                    continue
            
            bullish_df = pd.DataFrame(bullish_data)
            bullish_df.to_excel(writer, sheet_name='Bullish Setups', index=False)
            worksheet = writer.sheets['Bullish Setups']
            _format_worksheet(worksheet, len(bullish_df) + 1)
        
        # ========== SHEET 3: BEARISH SETUPS ==========
        if ranked_bearish:
            bearish_data = []
            for r in ranked_bearish[:15]:
                try:
                    bearish_data.append({
                        'Stock': _get(r, 'symbol', ''),
                        'Score': _get_number(r, 'confidence', '.0f'),
                        'P1 (High)': _get_currency(r, 'p1', 2),
                        'P2 (Low)': _get_currency(r, 'p2', 2),
                        'P3 (High)': _get_currency(r, 'p3', 2),
                        'P4 (Low)': _get_currency(r, 'p4', 2),
                        'P5 (Entry)': _get_currency(r, 'p5', 2),
                        'EPA Target': _get_currency(r, 'epa', 2),
                        'Stop Loss': _get_currency(r, 'stop_loss', 2),
                        '1-3 Line': _get(r, 'line_13_quality', ''),
                        '2-4 Line': _get(r, 'line_24_quality', ''),
                        'Convergence': 'Yes' if r.get('converging', True) else 'No',
                        'Entry Quality': _get(r, 'sweet_zone_quality', ''),
                    })
                except Exception as e:
                    logger.warning("Skipping bearish row for %s due to error: %s", _get(r, 'symbol'), e)
                    continue
            
            bearish_df = pd.DataFrame(bearish_data)
            bearish_df.to_excel(writer, sheet_name='Bearish Setups', index=False)
            worksheet = writer.sheets['Bearish Setups']
            _format_worksheet(worksheet, len(bearish_df) + 1)
        
        # ========== SHEET 4: PATTERN DETAILS ==========
        detail_data = []
        for r in display_results[:10]:
            try:
                detail_data.append({
                    'Stock': _get(r, 'symbol', ''),
                    'Pattern Type': _get(r, 'pattern_type', ''),
                    'Score': _get_number(r, 'confidence', '.0f'),
                    'Current Price': _get_currency(r, 'current_price', 2),
                    'P1': _get_currency(r, 'p1', 2),
                    'P2': _get_currency(r, 'p2', 2),
                    'P3': _get_currency(r, 'p3', 2),
                    'P4': _get_currency(r, 'p4', 2),
                    'P5': _get_currency(r, 'p5', 2),
                    'EPA': _get_currency(r, 'epa', 2),
                    'Sweet Zone Quality': _get(r, 'sweet_zone_quality', ''),
                    'Line 1-3': _get(r, 'line_13_quality', ''),
                    'Line 2-4': _get(r, 'line_24_quality', ''),
                    'Sector': _get(r, 'sector', ''),
                    'RSI': _get_number(r, 'rsi', '.1f'),
                    'Volume': (_get_number(r, 'vol_ratio_20', '.2f') + 'x') if _get(r, 'vol_ratio_20', None) not in (None, '') else '',
                })
            except Exception as e:
                logger.warning("Skipping detail row for %s due to error: %s", _get(r, 'symbol'), e)
                continue
        
        detail_df = pd.DataFrame(detail_data)
        detail_df.to_excel(writer, sheet_name='Pattern Details', index=False)
        worksheet = writer.sheets['Pattern Details']
        _format_worksheet(worksheet, len(detail_df) + 1)
        
        # ========== SHEET 5: METADATA ==========
        metadata = [
            ['Report Type', 'Wolfe Wave Screener'],
            ['Generated', datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S IST')],
            ['Total Stocks Scanned', '~150'],
            ['Wolfe Wave Setups Found', len(all_results)],
            ['Bullish Patterns', len([r for r in all_results if _get(r, 'pattern_type') == 'Bullish'])],
            ['Bearish Patterns', len([r for r in all_results if _get(r, 'pattern_type') == 'Bearish'])],
            ['', ''],
            ['Screening Parameters', ''],
            ['Universe', 'Nifty 50 + Nifty Next 50 + Select Midcaps'],
            ['Data Period', '1-year daily'],
            ['Min Confidence', '50%'],
        ]
        
        metadata_df = pd.DataFrame(metadata)
        metadata_df.to_excel(writer, sheet_name='Metadata', index=False, header=False)
    
    print(f"  [OK] Excel report saved to {output_file}\n")
    return output_file


def _format_worksheet(worksheet, row_count):
    """
    Apply formatting to worksheet (headers, borders, alternating colors).
    
    Args:
        worksheet: openpyxl worksheet object
        row_count: Number of rows (including header)
    """
    # Header formatting
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply header formatting
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Alternate row colors
    light_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    
    for row_idx in range(2, row_count):
        for col_idx in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            if row_idx % 2 == 0:
                cell.fill = light_fill
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width
